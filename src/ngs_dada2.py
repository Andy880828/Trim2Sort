"""NGS DADA2 分析模組

NGS DADA2 analysis module.
"""

import gc
from pathlib import Path
import re
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox

import customtkinter
import pandas as pd
from PIL import Image

from src.utils.logger_utils import get_logger
from src.utils.path_utils import (
    find_latest_ref_file,
    get_blastn_path,
    get_cutadapt_path,
    get_icon_path,
)
from src.utils.subprocess_utils import run_command
from src.utils.ui_config import COLORS, FONTS, LAYOUT

logger = get_logger(__name__)

# rpy2 將在需要時延遲導入, 避免在模組載入時就要求 R 環境
_ro = None
_importr = None


def _ensure_r_environment():
    """確保 R 環境已初始化並導入 rpy2

    Ensure R environment is initialized and import rpy2.

    Returns:
        tuple: (ro, importr) rpy2 模組 / (ro, importr) rpy2 modules.

    Raises:
        RuntimeError: 如果無法初始化 R 環境 / If R environment cannot be initialized.
    """
    global _ro, _importr

    if _ro is not None and _importr is not None:
        return _ro, _importr

    import os

    # 檢查並設定 R 環境
    from src.utils.path_utils import get_project_root

    project_root = get_project_root()
    renv_path = project_root / ".renv"

    # R 可能在 lib/R 或 Lib/R (micromamba 使用大寫 Lib)
    r_home = renv_path / "Lib" / "R"
    if not r_home.exists():
        r_home = renv_path / "lib" / "R"

    r_bin = renv_path / "bin"
    r_lib_bin = r_home / "bin" / "x64"
    library_bin = renv_path / "Library" / "bin"  # micromamba 依賴庫

    if not r_home.exists():
        msg = (
            f"R 環境不存在於 {r_home}。\n"
            f"請執行 execute(開啟請按我).bat 來建立 R 環境。\n\n"
            f"R environment not found at {r_home}.\n"
            f"Please run execute(開啟請按我).bat to create R environment."
        )
        raise RuntimeError(msg)

    # 設定 R_HOME
    if not os.environ.get("R_HOME"):
        os.environ["R_HOME"] = str(r_home)
        logger.info("設定 R_HOME=%s", r_home)

    # 將所有必要的 bin 目錄加入 PATH (重要: 這樣才能找到 R.dll 的依賴)
    current_path = os.environ.get("PATH", "")
    path_parts = [str(r_lib_bin), str(library_bin), str(r_bin)]
    # 確保這些路徑在 PATH 的最前面
    new_path = os.pathsep.join(path_parts) + os.pathsep + current_path
    os.environ["PATH"] = new_path
    logger.info("已將 R 路徑加入 PATH")

    # 使用 Windows API 設定 DLL 搜尋路徑 (Python 3.8+)
    try:
        if hasattr(os, "add_dll_directory"):
            os.add_dll_directory(str(library_bin))
            os.add_dll_directory(str(r_lib_bin))
            logger.debug("已使用 os.add_dll_directory 設定 DLL 搜尋路徑")
    except Exception:
        pass

    # 使用 ctypes 預先載入依賴 DLL, 確保載入順序正確
    try:
        import ctypes

        # 預先載入必要的依賴 DLL
        dll_names = [
            "libgcc_s_seh-1.dll",
            "libstdc++-6.dll",
            "libwinpthread-1.dll",
        ]
        for dll_name in dll_names:
            dll_path = library_bin / dll_name
            if dll_path.exists():
                try:
                    ctypes.CDLL(str(dll_path))
                    logger.debug("已預載入 %s", dll_name)
                except Exception:
                    pass  # 忽略載入失敗, 可能已經載入
    except Exception:
        pass  # ctypes 載入失敗不影響主流程

    try:
        import rpy2.robjects as ro
        from rpy2.robjects.packages import importr

        _ro = ro
        _importr = importr
        logger.info("R 環境初始化成功")
        return ro, importr
    except Exception as e:
        msg = f"無法初始化 R 環境: {e}\n請確認已執行 execute(開啟請按我).bat 來建立 R 環境。"
        raise RuntimeError(msg) from e


def load_app_image() -> customtkinter.CTkImage:
    """載入應用程式圖示

    Load application icon.

    Returns:
        customtkinter.CTkImage: 應用程式圖示 / Application icon.
    """
    icon_path = get_icon_path()
    return customtkinter.CTkImage(light_image=Image.open(icon_path), size=(128, 72))


class DADA2Config:
    """DADA2 分析設定

    DADA2 analysis configuration.
    """

    def __init__(self) -> None:
        """初始化 DADA2 設定

        Initialize DADA2 configuration.
        """
        self.primer_mode: str = "MiFish-U"
        self.forward_primer: str = "GTCGGTAAAACTCGTGCCAGC"
        self.reverse_primer: str = "CATAGTGGGGTATCTAATCCCAGTTTG"
        self.trunclen_forward: int = 160
        self.trunclen_reverse: int = 100
        self.max_ee_forward: int = 1
        self.max_ee_reverse: int = 1
        self.ref_path: str = find_latest_ref_file()


class ConfigWindow(customtkinter.CTkToplevel):
    """設定視窗

    Configuration window.
    """

    def __init__(self, parent: customtkinter.CTkFrame, *args, **kwargs) -> None:
        """初始化設定視窗

        Initialize configuration window.

        Args:
            parent: 父視窗 / Parent window.
            *args: 位置參數 / Positional arguments.
            **kwargs: 關鍵字參數 / Keyword arguments.
        """
        super().__init__(parent, *args, **kwargs)
        self.parent = parent
        self.title("DADA2 Configuration")
        self.geometry("480x580")
        self.configure(fg_color=COLORS.PRIMARY_BG)
        self.resizable(False, False)

        self._setup_ui()

    def _setup_ui(self) -> None:
        """設定使用者介面

        Setup user interface.
        """
        primer_label = customtkinter.CTkLabel(
            self,
            text="Primers:",
            text_color=COLORS.TEXT_PRIMARY,
            font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
        )
        primer_label.grid(row=0, column=0, padx=20, pady=(15, 3), sticky="w")

        self.primer_mode = tk.StringVar(value=self.parent.config.primer_mode)
        primer_combo = customtkinter.CTkComboBox(
            self,
            variable=self.primer_mode,
            values=["MiFish-U", "Manual"],
            width=370,
            height=LAYOUT.ENTRY_HEIGHT,
            border_width=LAYOUT.BORDER_WIDTH,
            corner_radius=LAYOUT.CORNER_RADIUS,
            fg_color=COLORS.SECONDARY_BG,
            text_color=COLORS.TEXT_PRIMARY,
            button_color=COLORS.HOVER_BG,
            button_hover_color=COLORS.BUTTON_HOVER,
            dropdown_fg_color=COLORS.SECONDARY_BG,
            dropdown_text_color=COLORS.TEXT_PRIMARY,
            dropdown_hover_color=COLORS.HOVER_BG,
            font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
            command=self.on_primer_mode_change,
        )
        primer_combo.grid(row=1, column=0, columnspan=2, padx=20, pady=(0, 5), sticky="w")

        forward_label = customtkinter.CTkLabel(
            self,
            text="Forward Primer:",
            text_color=COLORS.TEXT_PRIMARY,
            font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
        )
        forward_label.grid(row=2, column=0, padx=20, pady=(8, 3), sticky="w")

        self.forward_primer = tk.StringVar(value=self.parent.config.forward_primer)
        self.forward_entry = customtkinter.CTkEntry(
            self,
            textvariable=self.forward_primer,
            text_color=COLORS.TEXT_PRIMARY,
            width=440,
            height=LAYOUT.ENTRY_HEIGHT,
            border_width=LAYOUT.BORDER_WIDTH,
            corner_radius=LAYOUT.CORNER_RADIUS,
            fg_color=COLORS.SECONDARY_BG,
            font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
        )
        self.forward_entry.grid(row=3, column=0, columnspan=2, padx=20, pady=(0, 5), sticky="w")

        reverse_label = customtkinter.CTkLabel(
            self,
            text="Reverse Primer:",
            text_color=COLORS.TEXT_PRIMARY,
            font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
        )
        reverse_label.grid(row=4, column=0, padx=20, pady=(8, 3), sticky="w")

        self.reverse_primer = tk.StringVar(value=self.parent.config.reverse_primer)
        self.reverse_entry = customtkinter.CTkEntry(
            self,
            textvariable=self.reverse_primer,
            text_color=COLORS.TEXT_PRIMARY,
            width=440,
            height=LAYOUT.ENTRY_HEIGHT,
            border_width=LAYOUT.BORDER_WIDTH,
            corner_radius=LAYOUT.CORNER_RADIUS,
            fg_color=COLORS.SECONDARY_BG,
            font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
        )
        self.reverse_entry.grid(row=5, column=0, columnspan=2, padx=20, pady=(0, 5), sticky="w")

        trunclen_f_label = customtkinter.CTkLabel(
            self,
            text="Truncate Length (Forward):",
            text_color=COLORS.TEXT_PRIMARY,
            font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
        )
        trunclen_f_label.grid(row=6, column=0, padx=20, pady=(8, 3), sticky="w")

        self.trunclen_forward = tk.StringVar(value=str(self.parent.config.trunclen_forward))
        trunclen_f_entry = customtkinter.CTkEntry(
            self,
            textvariable=self.trunclen_forward,
            text_color=COLORS.TEXT_PRIMARY,
            width=440,
            height=LAYOUT.ENTRY_HEIGHT,
            border_width=LAYOUT.BORDER_WIDTH,
            corner_radius=LAYOUT.CORNER_RADIUS,
            fg_color=COLORS.SECONDARY_BG,
            font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
        )
        trunclen_f_entry.grid(row=7, column=0, columnspan=2, padx=20, pady=(0, 5), sticky="w")

        trunclen_r_label = customtkinter.CTkLabel(
            self,
            text="Truncate Length (Reverse):",
            text_color=COLORS.TEXT_PRIMARY,
            font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
        )
        trunclen_r_label.grid(row=8, column=0, padx=20, pady=(8, 3), sticky="w")

        self.trunclen_reverse = tk.StringVar(value=str(self.parent.config.trunclen_reverse))
        trunclen_r_entry = customtkinter.CTkEntry(
            self,
            textvariable=self.trunclen_reverse,
            text_color=COLORS.TEXT_PRIMARY,
            width=440,
            height=LAYOUT.ENTRY_HEIGHT,
            border_width=LAYOUT.BORDER_WIDTH,
            corner_radius=LAYOUT.CORNER_RADIUS,
            fg_color=COLORS.SECONDARY_BG,
            font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
        )
        trunclen_r_entry.grid(row=9, column=0, columnspan=2, padx=20, pady=(0, 5), sticky="w")

        maxee_f_label = customtkinter.CTkLabel(
            self,
            text="Max Expected Errors (Forward):",
            text_color=COLORS.TEXT_PRIMARY,
            font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
        )
        maxee_f_label.grid(row=10, column=0, padx=20, pady=(8, 3), sticky="w")

        self.max_ee_forward = tk.StringVar(value=str(self.parent.config.max_ee_forward))
        maxee_f_entry = customtkinter.CTkEntry(
            self,
            textvariable=self.max_ee_forward,
            text_color=COLORS.TEXT_PRIMARY,
            width=440,
            height=LAYOUT.ENTRY_HEIGHT,
            border_width=LAYOUT.BORDER_WIDTH,
            corner_radius=LAYOUT.CORNER_RADIUS,
            fg_color=COLORS.SECONDARY_BG,
            font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
        )
        maxee_f_entry.grid(row=11, column=0, columnspan=2, padx=20, pady=(0, 5), sticky="w")

        maxee_r_label = customtkinter.CTkLabel(
            self,
            text="Max Expected Errors (Reverse):",
            text_color=COLORS.TEXT_PRIMARY,
            font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
        )
        maxee_r_label.grid(row=12, column=0, padx=20, pady=(8, 3), sticky="w")

        self.max_ee_reverse = tk.StringVar(value=str(self.parent.config.max_ee_reverse))
        maxee_r_entry = customtkinter.CTkEntry(
            self,
            textvariable=self.max_ee_reverse,
            text_color=COLORS.TEXT_PRIMARY,
            width=440,
            height=LAYOUT.ENTRY_HEIGHT,
            border_width=LAYOUT.BORDER_WIDTH,
            corner_radius=LAYOUT.CORNER_RADIUS,
            fg_color=COLORS.SECONDARY_BG,
            font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
        )
        maxee_r_entry.grid(row=13, column=0, columnspan=2, padx=20, pady=(0, 5), sticky="w")

        ref_label = customtkinter.CTkLabel(
            self,
            text="Ref Path (Auto-detected):",
            text_color=COLORS.TEXT_PRIMARY,
            font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
        )
        ref_label.grid(row=14, column=0, padx=20, pady=(8, 3), sticky="w")

        self.ref_path = tk.StringVar(value=self.parent.config.ref_path)
        self.ref_entry = customtkinter.CTkEntry(
            self,
            textvariable=self.ref_path,
            text_color=COLORS.TEXT_PRIMARY,
            width=350,
            height=LAYOUT.ENTRY_HEIGHT,
            border_width=LAYOUT.BORDER_WIDTH,
            corner_radius=LAYOUT.CORNER_RADIUS,
            fg_color=COLORS.SECONDARY_BG,
            font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
        )
        self.ref_entry.grid(row=15, column=0, padx=20, pady=(0, 5), sticky="w")

        ref_button = customtkinter.CTkButton(
            self,
            width=70,
            height=LAYOUT.BUTTON_HEIGHT,
            border_width=0,
            corner_radius=LAYOUT.CORNER_RADIUS,
            text="SELECT",
            fg_color=COLORS.SECONDARY_BG,
            text_color=COLORS.TEXT_PRIMARY,
            hover_color=COLORS.HOVER_BG,
            font=(FONTS.FAMILY, FONTS.SIZE_NORMAL, FONTS.STYLE_BOLD),
            command=self.browse_ref,
        )
        ref_button.grid(row=15, column=1, padx=(0, 20), pady=(0, 5), sticky="w")

        button_frame = customtkinter.CTkFrame(self, fg_color="transparent")
        button_frame.grid(row=16, column=0, columnspan=2, padx=20, pady=(15, 15), sticky="ew")

        ok_button = customtkinter.CTkButton(
            button_frame,
            width=220,
            height=30,
            border_width=0,
            corner_radius=LAYOUT.CORNER_RADIUS,
            text="OK",
            fg_color=COLORS.SECONDARY_BG,
            text_color=COLORS.TEXT_PRIMARY,
            hover_color=COLORS.HOVER_BG,
            font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
            command=self.on_ok,
        )
        ok_button.pack(side="left", padx=(0, 10), expand=True, fill="x")

        cancel_button = customtkinter.CTkButton(
            button_frame,
            width=220,
            height=30,
            border_width=0,
            corner_radius=LAYOUT.CORNER_RADIUS,
            text="CANCEL",
            fg_color=COLORS.SECONDARY_BG,
            text_color=COLORS.TEXT_PRIMARY,
            hover_color=COLORS.HOVER_BG,
            font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
            command=self.on_cancel,
        )
        cancel_button.pack(side="left", expand=True, fill="x")

        self.on_primer_mode_change(self.primer_mode.get())

    def on_primer_mode_change(self, choice: str) -> None:
        """當 Primer 模式改變時調整欄位狀態

        Adjust field states when primer mode changes.

        Args:
            choice (str): 選擇的模式 / Selected mode.
        """
        match choice:
            case "MiFish-U":
                self.forward_primer.set("GTCGGTAAAACTCGTGCCAGC")
                self.reverse_primer.set("CATAGTGGGGTATCTAATCCCAGTTTG")
                self.forward_entry.configure(state="disabled")
                self.reverse_entry.configure(state="disabled")
            case "Manual":
                self.forward_entry.configure(state="normal")
                self.reverse_entry.configure(state="normal")

    def browse_ref(self) -> None:
        """瀏覽選擇 Ref 檔案

        Browse to select ref file.
        """
        filename = filedialog.askopenfilename(
            title="Select Reference File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if filename:
            self.ref_path.set(filename)

    def on_ok(self) -> None:
        """儲存配置並關閉視窗

        Save configuration and close window.
        """
        self.parent.config.primer_mode = self.primer_mode.get()
        self.parent.config.forward_primer = self.forward_primer.get()
        self.parent.config.reverse_primer = self.reverse_primer.get()

        try:
            self.parent.config.trunclen_forward = int(self.trunclen_forward.get())
        except ValueError:
            self.parent.config.trunclen_forward = 160
            logger.warning("Forward truncate length 設定無效, 使用預設值 160")

        try:
            self.parent.config.trunclen_reverse = int(self.trunclen_reverse.get())
        except ValueError:
            self.parent.config.trunclen_reverse = 100
            logger.warning("Reverse truncate length 設定無效, 使用預設值 100")

        try:
            self.parent.config.max_ee_forward = int(self.max_ee_forward.get())
        except ValueError:
            self.parent.config.max_ee_forward = 1
            logger.warning("Forward max EE 設定無效, 使用預設值 1")

        try:
            self.parent.config.max_ee_reverse = int(self.max_ee_reverse.get())
        except ValueError:
            self.parent.config.max_ee_reverse = 1
            logger.warning("Reverse max EE 設定無效, 使用預設值 1")

        self.parent.config.ref_path = self.ref_path.get()
        self.destroy()

    def on_cancel(self) -> None:
        """取消並關閉視窗

        Cancel and close window.
        """
        self.destroy()


class NGS_DADA2(customtkinter.CTkToplevel):
    """NGS DADA2 分析主視窗

    NGS DADA2 analysis main window.
    """

    def __init__(self, *args, **kwargs) -> None:
        """初始化 NGS DADA2 分析視窗

        Initialize NGS DADA2 analysis window.

        Args:
            *args: 位置參數 / Positional arguments.
            **kwargs: 關鍵字參數 / Keyword arguments.
        """
        super().__init__(*args, **kwargs)
        self.title("NGS Analysis - DADA2")
        self.geometry("500x550")
        self.configure(fg_color=COLORS.PRIMARY_BG)
        self.resizable(False, False)
        self.grid_rowconfigure(0, weight=0)
        self.grid_columnconfigure(0, weight=1)

        self.frame = DADA2HeaderFrame(master=self)
        self.frame.configure(fg_color=COLORS.PRIMARY_BG)
        self.frame.grid(row=1, column=0, padx=10, pady=(0, 15))

        self.frame_2 = DADA2ContentFrame(master=self)
        self.frame_2.configure(fg_color=COLORS.PRIMARY_BG, height=420)
        self.frame_2.grid(row=2, column=0, padx=20, pady=0, sticky="nsew")


class DADA2HeaderFrame(customtkinter.CTkFrame):
    """DADA2 標題框架

    DADA2 header frame.
    """

    def __init__(self, master: customtkinter.CTk | customtkinter.CTkFrame, **kwargs) -> None:
        """初始化標題框架

        Initialize header frame.

        Args:
            master: 父視窗 / Parent window.
            **kwargs: 關鍵字參數 / Keyword arguments.
        """
        super().__init__(master, **kwargs)

        my_image = load_app_image()

        logo = customtkinter.CTkLabel(self, image=my_image, text="")
        logo.grid(row=0, column=2, padx=10)

        buttons_config = [
            ("USER GUIDE", 0),
            ("DOCUMENTATION", 2),
            ("CONTACT US", 4),
        ]

        for text, col in buttons_config:
            button = customtkinter.CTkButton(
                self,
                width=100,
                height=LAYOUT.BUTTON_HEIGHT,
                border_width=0,
                text=text,
                fg_color="transparent",
                text_color=COLORS.TEXT_SECONDARY,
                hover_color=COLORS.SECONDARY_BG,
                font=(FONTS.FAMILY, FONTS.SIZE_NORMAL, FONTS.STYLE_BOLD),
            )
            button.grid(row=1, column=col, padx=10)

            if col < 4:
                sep_label = customtkinter.CTkLabel(
                    self,
                    text="✦",
                    text_color=COLORS.TEXT_SECONDARY,
                    font=(FONTS.FAMILY, 10, FONTS.STYLE_BOLD),
                )
                sep_label.grid(row=1, column=col + 1, padx=10)


class DADA2ContentFrame(customtkinter.CTkFrame):
    """DADA2 內容框架

    DADA2 content frame.
    """

    def __init__(self, master: customtkinter.CTk | customtkinter.CTkFrame, **kwargs) -> None:
        """初始化內容框架

        Initialize content frame.

        Args:
            master: 父視窗 / Parent window.
            **kwargs: 關鍵字參數 / Keyword arguments.
        """
        super().__init__(master, **kwargs)

        self.cutadapt_path = tk.StringVar()
        self.blastn_path = tk.StringVar()
        self.database_path = tk.StringVar()
        self.database_selector = tk.StringVar()
        self.samples_path = tk.StringVar()
        self.outputs_path = tk.StringVar()

        self.config = DADA2Config()

        self._setup_ui()
        self._auto_detect_tools()

    def _auto_detect_tools(self) -> None:
        """自動偵測工具路徑

        Auto-detect tool paths.
        """
        cutadapt_exe = get_cutadapt_path()
        blastn_exe = get_blastn_path()

        if cutadapt_exe.exists():
            self.cutadapt_path.set(str(cutadapt_exe))
        if blastn_exe.exists():
            self.blastn_path.set(str(blastn_exe))
        self._check_fields()

    def _setup_ui(self) -> None:
        """設定使用者介面

        Setup user interface.
        """
        cutadapt_info = tk.StringVar(
            value="Step1: Cutadapt.exe (Auto-detected from root directory)"
        )
        blastn_info = tk.StringVar(value="Step2: Blastn.exe (Auto-detected from root directory)")
        database_info = tk.StringVar(value="Step3: Select the folder directory of Database")
        samples_info = tk.StringVar(value="Step4: Select the folder directory of Samples")
        outputs_info = tk.StringVar(value="Step5: Select the folder directory of Outputs")
        instruction = tk.StringVar(value="Start the app after everthing above is selected")

        row = 0
        entries = [
            (cutadapt_info, self.cutadapt_path, self.browse_cutadapt),
            (blastn_info, self.blastn_path, self.browse_blastn),
        ]

        for info_var, path_var, command in entries:
            info_label = customtkinter.CTkLabel(
                self,
                textvariable=info_var,
                text_color=COLORS.TEXT_PRIMARY,
                font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
            )
            info_label.grid(row=row, column=0, padx=0, sticky="w")

            entry = customtkinter.CTkEntry(
                self,
                textvariable=path_var,
                text_color=COLORS.TEXT_PRIMARY,
                width=370,
                height=LAYOUT.ENTRY_HEIGHT,
                border_width=LAYOUT.BORDER_WIDTH,
                corner_radius=LAYOUT.CORNER_RADIUS,
                fg_color=COLORS.SECONDARY_BG,
                font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
            )
            entry.grid(row=row + 1, column=0, padx=0, pady=(0, 10), sticky="w")

            button = customtkinter.CTkButton(
                self,
                width=80,
                height=LAYOUT.BUTTON_HEIGHT,
                border_width=0,
                corner_radius=LAYOUT.CORNER_RADIUS,
                text="SELECT",
                fg_color=COLORS.SECONDARY_BG,
                text_color=COLORS.TEXT_PRIMARY,
                hover_color=COLORS.HOVER_BG,
                font=(FONTS.FAMILY, FONTS.SIZE_NORMAL, FONTS.STYLE_BOLD),
                command=command,
            )
            button.grid(row=row + 1, column=1, padx=(10, 0), pady=(0, 10), sticky="w")

            row += 2

        database_info_label = customtkinter.CTkLabel(
            self,
            textvariable=database_info,
            text_color=COLORS.TEXT_PRIMARY,
            font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
        )
        database_info_label.grid(row=row, column=0, padx=0, sticky="w")

        database_entry = customtkinter.CTkEntry(
            self,
            textvariable=self.database_path,
            text_color=COLORS.TEXT_PRIMARY,
            width=370,
            height=LAYOUT.ENTRY_HEIGHT,
            border_width=LAYOUT.BORDER_WIDTH,
            corner_radius=LAYOUT.CORNER_RADIUS,
            fg_color=COLORS.SECONDARY_BG,
            font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
        )
        database_entry.grid(row=row + 1, column=0, padx=0, pady=(0, 10), sticky="w")

        database_button = customtkinter.CTkButton(
            self,
            width=80,
            height=LAYOUT.BUTTON_HEIGHT,
            border_width=0,
            corner_radius=LAYOUT.CORNER_RADIUS,
            text="SELECT",
            fg_color=COLORS.SECONDARY_BG,
            text_color=COLORS.TEXT_PRIMARY,
            hover_color=COLORS.HOVER_BG,
            font=(FONTS.FAMILY, FONTS.SIZE_NORMAL, FONTS.STYLE_BOLD),
            command=self.browse_database,
        )
        database_button.grid(row=row + 1, column=1, padx=(10, 0), pady=(0, 10), sticky="w")

        row += 2

        database_selector_info = tk.StringVar(value="Select Database File (.nal/.ndb)")
        database_selector_info_label = customtkinter.CTkLabel(
            self,
            textvariable=database_selector_info,
            text_color=COLORS.TEXT_PRIMARY,
            font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
        )
        database_selector_info_label.grid(row=row, column=0, padx=0, sticky="w")

        self.database_selector_combo = customtkinter.CTkComboBox(
            self,
            variable=self.database_selector,
            values=[],
            width=370,
            height=LAYOUT.ENTRY_HEIGHT,
            border_width=LAYOUT.BORDER_WIDTH,
            corner_radius=LAYOUT.CORNER_RADIUS,
            fg_color=COLORS.SECONDARY_BG,
            text_color=COLORS.TEXT_PRIMARY,
            button_color=COLORS.HOVER_BG,
            button_hover_color=COLORS.BUTTON_HOVER,
            dropdown_fg_color=COLORS.SECONDARY_BG,
            dropdown_text_color=COLORS.TEXT_PRIMARY,
            dropdown_hover_color=COLORS.HOVER_BG,
            font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
        )
        self.database_selector_combo.grid(row=row + 1, column=0, padx=0, pady=(0, 10), sticky="w")

        row += 2

        folder_entries = [
            (samples_info, self.samples_path, self.browse_samples),
            (outputs_info, self.outputs_path, self.browse_outputs),
        ]

        for info_var, path_var, command in folder_entries:
            info_label = customtkinter.CTkLabel(
                self,
                textvariable=info_var,
                text_color=COLORS.TEXT_PRIMARY,
                font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
            )
            info_label.grid(row=row, column=0, padx=0, sticky="w")

            entry = customtkinter.CTkEntry(
                self,
                textvariable=path_var,
                text_color=COLORS.TEXT_PRIMARY,
                width=370,
                height=LAYOUT.ENTRY_HEIGHT,
                border_width=LAYOUT.BORDER_WIDTH,
                corner_radius=LAYOUT.CORNER_RADIUS,
                fg_color=COLORS.SECONDARY_BG,
                font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
            )
            entry.grid(row=row + 1, column=0, padx=0, pady=(0, 10), sticky="w")

            button = customtkinter.CTkButton(
                self,
                width=80,
                height=LAYOUT.BUTTON_HEIGHT,
                border_width=0,
                corner_radius=LAYOUT.CORNER_RADIUS,
                text="SELECT",
                fg_color=COLORS.SECONDARY_BG,
                text_color=COLORS.TEXT_PRIMARY,
                hover_color=COLORS.HOVER_BG,
                font=(FONTS.FAMILY, FONTS.SIZE_NORMAL, FONTS.STYLE_BOLD),
                command=command,
            )
            button.grid(row=row + 1, column=1, padx=(10, 0), pady=(0, 10), sticky="w")

            row += 2

        instruction_label = customtkinter.CTkLabel(
            self,
            textvariable=instruction,
            text_color=COLORS.TEXT_PRIMARY,
            font=(FONTS.FAMILY, FONTS.SIZE_MEDIUM, FONTS.STYLE_BOLD),
        )
        instruction_label.grid(row=row, column=0, columnspan=2, padx=0, pady=(10, 0), sticky="ew")

        row += 1

        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        config_button = customtkinter.CTkButton(
            self,
            height=LAYOUT.BUTTON_HEIGHT,
            border_width=0,
            corner_radius=LAYOUT.CORNER_RADIUS,
            text="CONFIG",
            fg_color=COLORS.SECONDARY_BG,
            text_color=COLORS.TEXT_PRIMARY,
            hover_color=COLORS.HOVER_BG,
            font=(FONTS.FAMILY, FONTS.SIZE_NORMAL, FONTS.STYLE_BOLD),
            command=self.open_config,
        )
        config_button.grid(row=row, column=0, padx=(0, 5), sticky="ew")

        self.analyse_button = customtkinter.CTkButton(
            self,
            height=LAYOUT.BUTTON_HEIGHT,
            border_width=0,
            corner_radius=LAYOUT.CORNER_RADIUS,
            text="ANALYSE",
            fg_color=COLORS.ACCENT,
            text_color=COLORS.PRIMARY_BG,
            hover_color=COLORS.ACCENT_HOVER,
            font=(FONTS.FAMILY, FONTS.SIZE_NORMAL, FONTS.STYLE_BOLD),
            command=self.analyse,
            state="disabled",
        )
        self.analyse_button.grid(row=row, column=1, padx=(5, 0), sticky="ew")

        self._setup_field_validation()

    def _setup_field_validation(self) -> None:
        """設定欄位驗證

        Setup field validation.
        """
        self.cutadapt_path.trace_add("write", lambda *_args: self._check_fields())
        self.blastn_path.trace_add("write", lambda *_args: self._check_fields())
        self.database_path.trace_add("write", lambda *_args: self._check_fields())
        self.database_selector.trace_add("write", lambda *_args: self._check_fields())
        self.samples_path.trace_add("write", lambda *_args: self._check_fields())
        self.outputs_path.trace_add("write", lambda *_args: self._check_fields())
        self._check_fields()

    def _check_fields(self) -> None:
        """檢查所有必要欄位是否已填寫

        Check if all required fields are filled.
        """
        all_filled = (
            bool(self.cutadapt_path.get())
            and bool(self.blastn_path.get())
            and bool(self.database_path.get())
            and bool(self.database_selector.get())
            and bool(self.samples_path.get())
            and bool(self.outputs_path.get())
        )

        if all_filled:
            self.analyse_button.configure(state="normal")
        else:
            self.analyse_button.configure(state="disabled")

    def open_config(self) -> None:
        """開啟配置視窗

        Open configuration window.
        """
        config_window = ConfigWindow(self)
        config_window.focus()

    def browse_cutadapt(self) -> None:
        """瀏覽選擇 Cutadapt 執行檔

        Browse to select Cutadapt executable.
        """
        filename = filedialog.askopenfilenames()
        if filename:
            self.cutadapt_path.set(filename[0] if isinstance(filename, tuple) else filename)
        self._check_fields()

    def browse_blastn(self) -> None:
        """瀏覽選擇 Blastn 執行檔

        Browse to select Blastn executable.
        """
        filename = filedialog.askopenfilenames()
        if filename:
            self.blastn_path.set(filename[0] if isinstance(filename, tuple) else filename)
        self._check_fields()

    def browse_database(self) -> None:
        """瀏覽選擇資料庫資料夾

        Browse to select database folder.
        """
        foldername = filedialog.askdirectory()
        if not foldername:
            return

        self.database_path.set(foldername)

        db_path = Path(foldername)
        taxdb_btd_path = db_path / "taxdb.btd"
        taxdb_bti_path = db_path / "taxdb.bti"

        if not taxdb_btd_path.exists() or not taxdb_bti_path.exists():
            messagebox.showwarning(
                "Warning", "no taxdb.btd & taxdb.bti files found, scientific name will be N/A"
            )

        database_files = []
        for file_path in db_path.iterdir():
            if file_path.is_file():
                if file_path.suffix == ".nal":
                    database_files.append(f"(Combined) {file_path.stem}")
                elif file_path.suffix == ".ndb":
                    file_name = file_path.name
                    if not re.search(r"\.\d{2}\.ndb$", file_name):
                        database_files.append(file_path.stem)

        if database_files:
            self.database_selector_combo.configure(values=database_files)
            self.database_selector.set(database_files[0])
        else:
            self.database_selector_combo.configure(values=[])
            self.database_selector.set("")
        self._check_fields()

    def browse_samples(self) -> None:
        """瀏覽選擇樣本資料夾

        Browse to select samples folder.
        """
        foldername = filedialog.askdirectory()
        if foldername:
            self.samples_path.set(foldername)
        self._check_fields()

    def browse_outputs(self) -> None:
        """瀏覽選擇輸出資料夾

        Browse to select outputs folder.
        """
        foldername = filedialog.askdirectory()
        if foldername:
            self.outputs_path.set(foldername)
        self._check_fields()

    def analyse(self) -> None:
        """執行 DADA2 分析

        Execute DADA2 analysis.
        """
        samples_dir = Path(self.samples_path.get())
        outputs_dir = Path(self.outputs_path.get())

        if not samples_dir.exists() or not outputs_dir.exists():
            logger.error("樣本或輸出資料夾不存在")
            messagebox.showerror("Error", "Samples or outputs folder does not exist")
            return

        # 檢查是否有 .R1.clean.fastq 和 .R2.clean.fastq 檔案
        r1_files = list(samples_dir.glob("*.R1.clean.fastq"))
        r2_files = list(samples_dir.glob("*.R2.clean.fastq"))

        if not r1_files or not r2_files:
            logger.error("未找到 .R1.clean.fastq 或 .R2.clean.fastq 檔案")
            messagebox.showerror(
                "Error",
                "No .R1.clean.fastq or .R2.clean.fastq files found.\n"
                "DADA2 requires paired-end reads with .R1.clean.fastq and .R2.clean.fastq extensions.",
            )
            return

        gc.collect()

        folder_names = [
            "A_cutadapt",
            "B_filtered",
            "C_dada2",
            "D_blasts",
            "E_final_report",
            "F_chinese_report",
        ]

        folders: dict[str, Path] = {}
        for folder_name in folder_names:
            folder_path = outputs_dir / folder_name
            if folder_path.exists():
                try:
                    shutil.rmtree(folder_path)
                except PermissionError:
                    for item in folder_path.iterdir():
                        try:
                            if item.is_dir():
                                shutil.rmtree(item, ignore_errors=True)
                            else:
                                item.unlink()
                        except Exception:
                            pass
            folder_path.mkdir(parents=True, exist_ok=True)
            folders[folder_name] = folder_path

        processor = DADA2Processor(
            cutadapt_path=self.cutadapt_path.get(),
            blastn_path=self.blastn_path.get(),
            database_path=self.database_path.get(),
            database_selector=self.database_selector.get(),
            config=self.config,
            folders=folders,
        )

        try:
            processor.run_analysis(samples_dir)
            logger.info("分析完成")
            messagebox.showinfo("Success", "Analysis completed successfully")
            # Cheers! 原作者的願望
            processor.show_cheers()
        except Exception as e:
            logger.error(f"分析過程發生錯誤: {e}")
            messagebox.showerror("Error", f"Analysis failed: {e}")


class DADA2Processor:
    """DADA2 分析處理器

    DADA2 analysis processor.
    """

    def __init__(
        self,
        cutadapt_path: str,
        blastn_path: str,
        database_path: str,
        database_selector: str,
        config: DADA2Config,
        folders: dict[str, Path],
    ) -> None:
        """初始化 DADA2 處理器

        Initialize DADA2 processor.

        Args:
            cutadapt_path (str): Cutadapt 執行檔路徑 / Cutadapt executable path.
            blastn_path (str): Blastn 執行檔路徑 / Blastn executable path.
            database_path (str): 資料庫路徑 / Database path.
            database_selector (str): 資料庫選擇器 / Database selector.
            config (DADA2Config): DADA2 設定 / DADA2 configuration.
            folders (dict[str, Path]): 資料夾路徑字典 / Folder paths dictionary.
        """
        self.cutadapt_path = cutadapt_path
        self.blastn_path = blastn_path
        self.database_path = database_path
        self.database_selector = database_selector
        self.config = config
        self.folders = folders

    def run_analysis(self, samples_dir: Path) -> None:
        """執行完整分析流程

        Execute complete analysis workflow.

        Args:
            samples_dir (Path): 樣本資料夾路徑 / Samples directory path.
        """
        logger.info("開始 DADA2 分析")

        # 步驟 1: 檢查並安裝 DADA2
        logger.info("步驟 1/7: 檢查 R 環境與 DADA2 套件")
        self._check_and_install_dada2()

        # 步驟 2: 執行 DADA2 分析 (cutadapt, filter, denoise, merge, remove chimeras)
        logger.info("步驟 2/7: 執行 DADA2 主要分析流程")
        self._run_dada2_pipeline(samples_dir)

        # 步驟 3: 取得樣本列表
        dada2_dir = self.folders["C_dada2"]
        fasta_files = list(dada2_dir.glob("*_dada2.fasta"))
        # 使用正則取 _ 分割的第一個部分作為樣本名稱
        sample_names = []
        for f in fasta_files:
            stem = f.stem.replace("_dada2", "")
            # 取第一個 _ 之前的部分作為樣本名稱
            match = re.match(r"^([^_]+)", stem)
            if match:
                sample_names.append(match.group(1))
            else:
                sample_names.append(stem)

        if not sample_names:
            logger.warning("未找到任何 DADA2 輸出檔案")
            return

        logger.info(f"找到 {len(sample_names)} 個樣本")

        # 步驟 4: 執行 BLAST
        logger.info(f"步驟 3/7: 執行 BLAST (共 {len(sample_names)} 個樣本)")
        for sample_name in sample_names:
            self._run_blast(sample_name)
        gc.collect()

        # 步驟 5: 合併 reads 與 BLAST 結果
        logger.info(f"步驟 4/7: 合併 reads 與 BLAST 結果 (共 {len(sample_names)} 個樣本)")
        for sample_name in sample_names:
            self._merge_reads_and_blast(sample_name)
        gc.collect()

        # 步驟 6: 過濾重複資料
        logger.info(f"步驟 5/7: 過濾重複資料 (共 {len(sample_names)} 個樣本)")
        for sample_name in sample_names:
            self._filter_duplicates(sample_name)
        gc.collect()

        # 步驟 7: 加入中文名稱
        logger.info(f"步驟 6/7: 加入中文名稱 (共 {len(sample_names)} 個樣本)")
        for sample_name in sample_names:
            self._add_chinese_names(sample_name)
        gc.collect()

        logger.info("步驟 7/7: DADA2 分析完成")

    def _check_and_install_dada2(self) -> None:
        """檢查並安裝 DADA2 套件

        Check and install DADA2 package.
        """
        try:
            ro, importr = _ensure_r_environment()
            logger.info("檢查 DADA2 套件是否已安裝...")

            # 嘗試載入 DADA2
            try:
                importr("dada2")
                logger.info("DADA2 套件已安裝")
                return
            except Exception:
                logger.info("DADA2 套件未安裝, 開始安裝...")

            # 安裝 BiocManager (如果未安裝)
            ro.r(
                """
            if (!requireNamespace("BiocManager", quietly = TRUE)) {
                install.packages("BiocManager", repos = "https://cloud.r-project.org/")
            }
            """
            )

            # 安裝 DADA2 (讓 BiocManager 自動選擇適合當前 R 版本的版本)
            ro.r('BiocManager::install("dada2", ask = FALSE)')

            logger.info("DADA2 套件安裝完成")

        except Exception as e:
            logger.error(f"安裝 DADA2 套件失敗: {e}")
            raise RuntimeError(f"Failed to install DADA2: {e}") from e

    def _run_dada2_pipeline(self, samples_dir: Path) -> None:
        """執行 DADA2 分析管線

        Run DADA2 analysis pipeline.

        Args:
            samples_dir (Path): 樣本資料夾路徑 / Samples directory path.
        """
        try:
            ro, importr = _ensure_r_environment()
            logger.info("載入 DADA2 與相關套件...")
            dada2 = importr("dada2")
            importr("ShortRead")
            importr("Biostrings")

            # 取得檔案列表
            r_samples_dir = str(samples_dir).replace("\\", "/")

            # 尋找配對檔案
            fnFs = ro.r(
                f'sort(list.files("{r_samples_dir}", pattern=".R1.clean.fastq", full.names = TRUE))'
            )
            fnRs = ro.r(
                f'sort(list.files("{r_samples_dir}", pattern=".R2.clean.fastq", full.names = TRUE))'
            )

            # 取得樣本名稱
            ro.r.assign("fnFs", fnFs)
            sample_names = ro.r('sapply(strsplit(basename(fnFs), ".R1.clean.fastq"), `[`, 1)')
            sample_cnt = len(sample_names)
            logger.info(f"找到 {sample_cnt} 個樣本")

            # 設定 cutadapt 路徑與 primers
            cutadapt_path_r = str(Path(self.cutadapt_path)).replace("\\", "/")
            fwd = self.config.forward_primer
            rev = self.config.reverse_primer

            # Cutadapt 修剪 primers
            logger.info("執行 Cutadapt 修剪 primers...")
            path_cut = self.folders["A_cutadapt"]
            path_cut_r = str(path_cut).replace("\\", "/")

            ro.r.assign("fnFs", fnFs)
            ro.r.assign("fnRs", fnRs)
            fnFs_cut = ro.r(f'file.path("{path_cut_r}", basename(fnFs))')
            fnRs_cut = ro.r(f'file.path("{path_cut_r}", basename(fnRs))')

            # 取得反向互補序列
            ro.r.assign("FWD", fwd)
            ro.r.assign("REV", rev)
            ro.r("FWD.RC <- dada2:::rc(FWD)")
            ro.r("REV.RC <- dada2:::rc(REV)")

            # 執行 cutadapt
            ro.r.assign("cutadapt", cutadapt_path_r)
            ro.r.assign("fnFs_cut", fnFs_cut)
            ro.r.assign("fnRs_cut", fnRs_cut)
            ro.r(
                """
            R1.flags <- paste("-g", FWD, "-a", REV.RC)
            R2.flags <- paste("-G", REV, "-A", FWD.RC)
            for(i in seq_along(fnFs)) {
                system2(cutadapt, args = c(R1.flags, R2.flags, "-n", 2,
                                         "-o", fnFs_cut[i], "-p", fnRs_cut[i],
                                         fnFs[i], fnRs[i]),
                       stdout = FALSE, stderr = FALSE)
            }
            """
            )

            # 過濾與修剪
            logger.info("執行品質過濾...")
            filtered_path = self.folders["B_filtered"]
            filtered_path_r = str(filtered_path).replace("\\", "/")

            fnFcs = ro.r(
                f'sort(list.files("{path_cut_r}", pattern=".R1.clean.fastq", full.names = TRUE))'
            )
            fnRcs = ro.r(
                f'sort(list.files("{path_cut_r}", pattern=".R2.clean.fastq", full.names = TRUE))'
            )

            ro.r.assign("sample_names", sample_names)
            filtFs = ro.r(
                f'file.path("{filtered_path_r}", paste0(sample_names, "_F_filt.fastq.gz"))'
            )
            filtRs = ro.r(
                f'file.path("{filtered_path_r}", paste0(sample_names, "_R_filt.fastq.gz"))'
            )

            ro.r.assign("filtFs", filtFs)
            ro.r.assign("filtRs", filtRs)
            ro.r("names(filtFs) <- sample_names")
            ro.r("names(filtRs) <- sample_names")
            filtFs = ro.r("filtFs")
            filtRs = ro.r("filtRs")

            trunclen_f = self.config.trunclen_forward
            trunclen_r = self.config.trunclen_reverse
            maxee_f = self.config.max_ee_forward
            maxee_r = self.config.max_ee_reverse

            dada2.filterAndTrim(
                fnFcs,
                filtFs,
                fnRcs,
                filtRs,
                truncLen=ro.IntVector([trunclen_f, trunclen_r]),
                maxN=0,
                maxEE=ro.IntVector([maxee_f, maxee_r]),
                truncQ=2,
                rm_phix=True,
                compress=True,
                multithread=False,
            )

            # 學習錯誤率
            logger.info("學習錯誤率...")
            errF = dada2.learnErrors(filtFs, multithread=True)
            errR = dada2.learnErrors(filtRs, multithread=True)

            # 核心演算法
            logger.info("執行核心樣本推論演算法...")
            dadaFs = dada2.dada(filtFs, err=errF, multithread=True)
            dadaRs = dada2.dada(filtRs, err=errR, multithread=True)

            # 合併配對
            logger.info("合併配對序列...")
            mergers = dada2.mergePairs(dadaFs, filtFs, dadaRs, filtRs, verbose=True)

            # 建立序列表
            logger.info("建立序列表...")
            seqtab = dada2.makeSequenceTable(mergers)

            # 移除嵌合體
            logger.info("移除嵌合體...")
            seqtab_nochim = dada2.removeBimeraDenovo(
                seqtab, method="consensus", multithread=True, verbose=True
            )

            # 輸出每個樣本的 fasta 與 track csv
            logger.info("輸出 DADA2 結果...")
            dada2_path = self.folders["C_dada2"]
            dada2_path_r = str(dada2_path).replace("\\", "/")

            ro.r.assign("seqtab_nochim", seqtab_nochim)
            ro.r.assign("dada2path", dada2_path_r)
            ro.r(
                """
            for (i in 1:nrow(seqtab_nochim)) {
                # 取得原始樣本名稱並取第一個 _ 之前的部分
                raw_name <- row.names(seqtab_nochim)[i]
                raw_name_char <- tryCatch({
                    as.character(raw_name)
                }, error = function(e) {
                    paste0("sample_", i)
                })
                # 嘗試分割並取第一個部分
                if (!is.na(raw_name_char) && nchar(raw_name_char) > 0 && grepl("_", raw_name_char)) {
                    sample_name_parts <- tryCatch({
                        strsplit(raw_name_char, "_", fixed=TRUE)[[1]]
                    }, error = function(e) {
                        raw_name_char
                    })
                    if (is.character(sample_name_parts) && length(sample_name_parts) > 0 && nchar(sample_name_parts[1]) > 0) {
                        clean_sample_name <- sample_name_parts[1]
                    } else {
                        clean_sample_name <- raw_name_char
                    }
                } else {
                    clean_sample_name <- raw_name_char
                }
                dada2name <- paste0(dada2path, "/", clean_sample_name, "_dada2.fasta")
                dada2trackname <- paste0(dada2path, "/", clean_sample_name, "_dada2track.csv")

                nonzero_col <- which(seqtab_nochim[i,] != 0)
                file_out <- file(dada2name, open = "a")

                if (length(nonzero_col) > 0) {
                    cntc <- 0
                    col_names <- colnames(seqtab_nochim)[nonzero_col]
                    for (col_name in col_names) {
                        cntc = cntc + 1
                        writeLines(paste0('>ASV',cntc,'\\n',col_name,'\\n'), file_out)
                    }
                    cntr <- length(nonzero_col)
                    readnumlist <- data.frame(ASV = numeric(cntr), Value = character(cntr))
                    readnums <- seqtab_nochim[i,nonzero_col]
                    total_sum <- sum(unlist(readnums))
                    ratio <- readnums / total_sum
                    readnumlist$ASV <- paste("ASV", 1:cntr, sep="")
                    readnumlist$Value <- unlist(readnums)
                    readnumlist$Ratio <- unlist(ratio)
                    write.csv(readnumlist, file = dada2trackname, row.names = FALSE)
                }
                close(file_out)
            }
            """
            )

            logger.info("DADA2 分析管線完成")

        except Exception as e:
            logger.error(f"DADA2 分析管線執行失敗: {e}")
            raise RuntimeError(f"DADA2 pipeline failed: {e}") from e

    def _run_blast(self, sample_name: str) -> None:
        """執行 BLAST

        Run BLAST.

        Args:
            sample_name (str): 樣本名稱 / Sample name.
        """
        logger.info(f"執行 BLAST: {sample_name}")
        dada2_dir = self.folders["C_dada2"]
        blast_dir = self.folders["D_blasts"]

        fasta_file = dada2_dir / f"{sample_name}_dada2.fasta"
        blast_output = blast_dir / f"{sample_name}_blasted.txt"

        db_display_name = self.database_selector
        if db_display_name.startswith("(Combined) "):
            db_name = db_display_name.replace("(Combined) ", "", 1)
        else:
            db_name = db_display_name

        blast_cmd = [
            self.blastn_path,
            "-query",
            str(fasta_file),
            "-out",
            str(blast_output),
            "-db",
            db_name,
            "-outfmt",
            "6 qseqid pident qcovs sscinames sallacc",
            "-max_target_seqs",
            "2",
        ]
        run_command(blast_cmd, cwd=self.database_path)
        logger.info(f"完成執行 BLAST: {sample_name}")

    def _merge_reads_and_blast(self, sample_name: str) -> None:
        """合併 reads 與 BLAST 結果

        Merge reads and BLAST results.

        Args:
            sample_name (str): 樣本名稱 / Sample name.
        """
        try:
            logger.info(f"合併 reads 與 BLAST 結果: {sample_name}")
            dada2_dir = self.folders["C_dada2"]
            blast_dir = self.folders["D_blasts"]
            report_dir = self.folders["E_final_report"]

            track_file = dada2_dir / f"{sample_name}_dada2track.csv"
            blast_file = blast_dir / f"{sample_name}_blasted.txt"
            output_file = report_dir / f"{sample_name}_report.csv"

            # 讀取 track 資料
            track_data = pd.read_csv(track_file)

            # 讀取 blast 資料
            blast_data = pd.read_csv(
                blast_file,
                sep="\t",
                header=None,
                names=["ASV", "Identity", "Coverage", "Scientific_name", "Accession_number"],
            )

            # 合併資料
            track_data.set_index("ASV", inplace=True)
            blast_data.set_index("ASV", inplace=True)
            merged_data = pd.merge(
                track_data, blast_data, left_index=True, right_index=True, how="outer"
            )

            # 排序
            def extract_number(string: str) -> int:
                """提取字串中的數字

                Extract number from string.

                Args:
                    string (str): 輸入字串 / Input string.

                Returns:
                    int: 提取的數字 / Extracted number.
                """
                matches = re.findall(r"\d+", string)
                return int(matches[0]) if matches else 0

            merged_data = merged_data.iloc[merged_data.index.map(extract_number).argsort()]
            merged_data.to_csv(output_file, index=True, index_label="ASV")

            logger.info(f"完成合併: {sample_name}")

        except Exception as e:
            logger.error(f"合併失敗 {sample_name}: {e}")

    def _filter_duplicates(self, sample_name: str) -> None:
        """過濾重複資料

        Filter duplicate data.

        Args:
            sample_name (str): 樣本名稱 / Sample name.
        """
        try:
            logger.info(f"過濾重複資料: {sample_name}")
            report_dir = self.folders["E_final_report"]
            report_file = report_dir / f"{sample_name}_report.csv"

            origin_report = pd.read_csv(report_file, header=0)
            drop_duplicate = origin_report.drop_duplicates(
                subset=["ASV", "Scientific_name"], keep="first"
            ).copy()

            drop_duplicate.loc[:, "Identity"] = drop_duplicate["Identity"].astype(float)
            max_identities = drop_duplicate.groupby(["ASV"])["Identity"].transform("max")
            # 明確複製以避免 SettingWithCopyWarning
            filtered_data = drop_duplicate[drop_duplicate["Identity"] == max_identities].copy()

            # 轉換欄位類型為 object 以避免 dtype 警告
            if "Value" in filtered_data.columns:
                filtered_data.loc[:, "Value"] = filtered_data["Value"].astype(object)
            if "Ratio" in filtered_data.columns:
                filtered_data.loc[:, "Ratio"] = filtered_data["Ratio"].astype(object)

            # 設定重複項目的 Value 和 Ratio 為空字串
            duplicate_mask = filtered_data.duplicated(subset=["ASV"], keep="first")
            filtered_data.loc[duplicate_mask, "Value"] = ""
            filtered_data.loc[duplicate_mask, "Ratio"] = ""

            filtered_data.to_csv(report_file, index=False)
            logger.info(f"完成過濾: {sample_name}")

        except Exception as e:
            logger.error(f"過濾失敗 {sample_name}: {e}")

    def _add_chinese_names(self, sample_name: str) -> None:
        """加入中文名稱

        Add Chinese names.

        Args:
            sample_name (str): 樣本名稱 / Sample name.
        """
        try:
            logger.info(f"加入中文名稱: {sample_name}")
            report_dir = self.folders["E_final_report"]
            chinese_dir = self.folders["F_chinese_report"]

            report_file = report_dir / f"{sample_name}_report.csv"
            output_file = chinese_dir / f"{sample_name}_chreport.xlsx"

            ref_path = self.config.ref_path
            if not ref_path:
                logger.warning("未找到參考檔案")
                return

            ref_data = pd.read_excel(ref_path, engine="openpyxl")
            report_data = pd.read_csv(report_file)

            ref_data.set_index("Scientific_name", inplace=True)
            report_data.set_index("Scientific_name", inplace=True)

            merged_data = pd.merge(
                report_data, ref_data, left_index=True, right_index=True, how="left"
            )
            merged_data.insert(5, "Scientific_name", merged_data.index)

            # 排序
            def extract_number(string: str) -> int:
                """提取字串中的數字

                Extract number from string.

                Args:
                    string (str): 輸入字串 / Input string.

                Returns:
                    int: 提取的數字 / Extracted number.
                """
                matches = re.findall(r"\d+", str(string))
                return int(matches[0]) if matches else 0

            merged_data = merged_data.sort_values(by="ASV", key=lambda x: x.map(extract_number))

            # 先保存為 Excel, 再套用樣式
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill

            # 先保存基本資料
            merged_data.to_excel(output_file, index=False, sheet_name="Sheet1", engine="openpyxl")

            # 載入並套用樣式
            wb = load_workbook(output_file)
            ws = wb["Sheet1"]

            # 確保工作表可見
            ws.sheet_state = "visible"

            # 套用樣式: 根據 Stat 欄位標記顏色
            if "Stat" in merged_data.columns:
                stat_col_idx = merged_data.columns.get_loc("Stat") + 1
                dupe_fill = PatternFill(start_color="00c2c7", end_color="00c2c7", fill_type="solid")
                filtered_fill = PatternFill(
                    start_color="94F7B2", end_color="94F7B2", fill_type="solid"
                )

                for row_idx in range(2, ws.max_row + 1):
                    stat_value = ws.cell(row=row_idx, column=stat_col_idx).value
                    if stat_value == "*DUPE*":
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = dupe_fill
                    elif stat_value == "*FILTERED*":
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = filtered_fill

            wb.save(output_file)
            wb.close()

            logger.info(f"完成加入中文名稱: {sample_name}")

        except Exception as e:
            logger.error(f"加入中文名稱失敗 {sample_name}: {e}")

    def show_cheers(self) -> None:
        """顯示慶祝動畫(by逸晴)

        Show celebration animation(by Yiggi).
        """
        try:
            ro, _ = _ensure_r_environment()
            logger.info("Cheers! 分析完成!")
            ro.r(
                """
            library(grDevices)
            x11() # Open graphics device
            plot(0, 0, type="n", xlim=c(-2,2), ylim=c(-2,2), xlab="", ylab="", main="Analysis Complete!")

            # Draw a star
            t <- seq(0, 2*pi, length.out=11)
            r <- rep(c(1, 0.4), length.out=11)
            x <- r * cos(t)
            y <- r * sin(t)
            polygon(x, y, col="yellow", border="red", lwd=2)
            text(0, 0, "SUCCESS!", cex=2, col="red", font=2)

            Sys.sleep(3)
            dev.off()
            """
            )
        except Exception as e:
            logger.warning(f"無法顯示慶祝動畫: {e}")
            # 不影響主流程
